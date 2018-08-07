#!/usr/bin/python
import sys
<<<<<<< HEAD
import os
import time
import os.path
from tempfile import mkstemp
from shutil import move
from os import remove, close
from subprocess import call # This is needed to submit jobs

"""
This code is used to submit mcce jobs since we have 50 systems 
"""
 
top_location = '/home/salah/ryr1_energyCalculations/calculations/5t9r/'

def myPlot(ch_energy_list,occ_list):
    
    plt.plot(ch_energy_list,occ_list, 'ro')
    #plt.axis([1, 50, 10.5, 12.5])
    plt.show()

occ_list = []
ch_energy_list = []


for i in range(1,51):
	# get fort.38
	fort38 = top_location + "frame_" +str(i).zfill(2)+'/binding/Ca/100_0/fort.38'
	fort38_lines = open(filename).readlines() 
	for line in fort38_lines:
		line =line.split()
		if line[0] == 'extra':
			for fields in line:
				ch_energy_list.append(fields)
		if line[0] == '_CADMB5102_002':
			for fields in line[1:]:
				occ_list.append(fields)
					


with open("data.txt",'w') as energies:
	for occ_list in nsmallest(1, list_of_occ, key=lambda x: abs(x-0.5)):
		
		if float(occ) == occ_list:
			energies.write("frame_%02d" % (i,)+'\t'+extra[kk+1]+'\t'+str(occ_list)+'\n')
	energies.close()











	

=======
import re
import math
import datetime
import os
import re
import openpyxl
from openpyxl import Workbook
from heapq import nsmallest


dirname = '/home/guest/ryr1_energyCalculations/calculations/5t9r/'


wb = openpyxl.Workbook()
ws3 = wb.create_sheet(index=0, title="overview")
ws3 = wb.create_sheet(index=1, title="binding_energies")

cell_header1 = 'B1' 
ws3[cell_header1] = "50/50 occupation energy (kCal/mol)"

all_binding_energies = []
with open("data.txt",'a') as energies:
        energies.write("July 10th, 2018.\nThis file contains the binding energies for each frame\n")
	for i in range(1,51):
		frame = dirname+"frame_" +str(i).zfill(2)+'/binding/Ca/100_0/fort.38'
		if not os.path.exists(frame):
			print dirname+"frame_" +str(i).zfill(2)
			continue
			
		sheetName = "frame_%02d" % (i,)
		pdb_lines = open(frame).readlines() 
		#print (frame)
		# create sheet
		ws3 = wb.create_sheet(index=i+1, title=sheetName)
		
		# insert sheet header
		ws3_sheet = wb[sheetName]
		cell_header1 = 'A1' 
		ws3_sheet[cell_header1] = 'chemical energy (kCal/mol)'
		cell_header1 = 'B1' 
		ws3_sheet[cell_header1] = 'occupation'
		# insert data into sheet
		j = 0
		k = 0
		kk = 0
		list_of_occ = []
		for line in pdb_lines:
			fields = line.split()
			if fields[0] == 'extra':
				col = line.split()
				list_of_ene = []
				extra = col
				for item in col[1:]: 
					cell_header = 'A'+str(j+2)
					ws3_sheet[cell_header] = item
					list_of_ene.append(float(item))
					j = j + 1
			if fields[0] == '_CADMB5102_002':
				col = line.split()
				for item in col[1:]:
					list_of_occ.append(float(item))
					cell_header = 'B1'
					ws3_sheet[cell_header] = fields[0]
					cell_header = 'B'+str(k+2)	
					ws3_sheet[cell_header] = item
					#energies.write(frame+'\t')
					k = k + 1
			print (nsmallest(1, list_of_occ, key=lambda x: abs(x-0.5)))
			if fields[0] == '_CADMB5102_002':
				col = line.split()
				for item in col[1:]:
					print (item)
					for occ_list in nsmallest(1, list_of_occ, key=lambda x: abs(x-0.5)):
						if float(item) == occ_list:
							energies.write("frame_%02d" % (i,)+'\t'+extra[kk+1]+'\t'+str(occ_list)+'\n')
					kk = kk + 1
				
					
		
		wb.save('data.xlsx')
	energies.close()
print ('Done')


>>>>>>> ab305d0df641f64ed48878f7bd344e7668d7e8c3
