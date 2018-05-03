#!/usr/bin/python
import sys
import re
import math
import datetime
import os
import re
import openpyxl
from openpyxl import Workbook
from heapq import nsmallest


dirname = "./frame_"


wb = openpyxl.Workbook()
ws3 = wb.create_sheet(index=0, title="overview")
ws3 = wb.create_sheet(index=1, title="binding_energies")

cell_header1 = 'B1' 
ws3[cell_header1] = "50/50 occupation energy (kCal/mol)"

all_binding_energies = []
with open("data.txt",'w') as energies:
	energies.write("This file contains the binding energies for each frame\n")
	for i in range(1,51):
		frame = dirname+"%02d" % (i,)+"/binding/fort.38"
		sheetName = "frame_%02d" % (i,)
		pdb_lines = open(frame).readlines() 
		print (frame)
		# create sheet
		ws3 = wb.create_sheet(index=i+1, title=sheetName)
		
		# insert sheet header
		ws3_sheet = wb[sheetName]
		cell_header1 = 'A1' 
		ws3_sheet[cell_header1] = 'chemical energy (kCal/mol)'
		cell_header1 = 'B1' 
		ws3_sheet[cell_header1] = 'occupation'
		# insert data into sheet
		j = 0
		k = 0
		kk = 0
		list_of_occ = []
		for line in pdb_lines:
			fields = line.split()
			if fields[0] == 'extra':
				col = line.split()
				list_of_ene = []
				extra = col
				for item in col[1:]: 
					cell_header = 'A'+str(j+2)
					ws3_sheet[cell_header] = item
					list_of_ene.append(float(item))
					j = j + 1
			if fields[0] == '_CADMG5104_002':
				col = line.split()
				for item in col[1:]:
					list_of_occ.append(float(item))
					cell_header = 'B1'
					ws3_sheet[cell_header] = fields[0]
					cell_header = 'B'+str(k+2)	
					ws3_sheet[cell_header] = item
					#energies.write(frame+'\t')
					k = k + 1
			#print (type(nsmallest(1, list_of_occ, key=lambda x: abs(x-0.5))))
			if fields[0] == '_CADMG5104_002':
				col = line.split()
				for item in col[1:]:
					for occ_list in nsmallest(1, list_of_occ, key=lambda x: abs(x-0.5)):
						if float(item) == occ_list:
							energies.write("frame_%02d" % (i,)+'\t'+extra[kk+1]+'\t'+str(occ_list)+'\n')
					kk = kk + 1
				
					
		
		wb.save('data.xlsx')
		""" 
		with open(new_file,'w') as fp:
	    		fp.write('REMARK: This file was greated by getI_J.py that cut the I and G subunits from ' + filename + ' and saved into ' + new_file + '\n')
	    		for line in pdb_lines:
	        		fields = line.split()
	        		if len(fields) > 8:
					if fields[3] == 'I' or fields[3] == 'J':
	                		fp.write(line)
	    	fp.close()"""
	energies.close()
print ('Done')

